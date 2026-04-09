#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from flask import Flask, request, jsonify, send_file
from openpyxl import load_workbook
import json
import io
import os

app = Flask(__name__)

LANGS = ["DE", "EN", "ES", "IT", "NL", "PT"]

@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})

@app.route("/inject", methods=["POST"])
def inject():
    try:
        if "fichier" not in request.files:
            return jsonify({"error": "Fichier manquant"}), 400

        fichier = request.files["fichier"]
        traductions_raw = request.form.get("traductions", "[]")
        traductions = json.loads(traductions_raw)

        index = {}
        for row in traductions:
            fr = str(row.get("FR", "")).strip()
            if fr:
                index[fr] = row

        wb = load_workbook(fichier)
        ws = wb.active

        headers = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=c).value
            if val:
                headers[val] = c

        if "FR" not in headers:
            return jsonify({"error": "Colonne FR introuvable"}), 400

        fr_col = headers["FR"]
        injected = 0

        for r in range(2, ws.max_row + 1):
            fr_val = ws.cell(row=r, column=fr_col).value
            if not fr_val or str(fr_val).strip() == "":
                continue

            fr_text = " ".join(str(fr_val).split()).strip()
            row_data = index.get(fr_text)
            if not row_data:
                continue

            for lg in LANGS:
                if lg not in headers:
                    continue
                cell = ws.cell(row=r, column=headers[lg])
                if cell.value is None or str(cell.value).strip() == "":
                    val = str(row_data.get(lg, "")).strip()
                    if val:
                        cell.value = val
                        injected += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        nom_original = fichier.filename or "fichier.xlsx"
        nom_sortie = nom_original.replace(".xlsx", "-traduit.xlsx")

        print(f"✅ {injected} cellules injectées dans {nom_sortie}", flush=True)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=nom_sortie
        )

    except Exception as e:
        print(f"❌ Erreur: {e}", flush=True)
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    print(f"🚀 Serveur démarré sur http://localhost:{port}", flush=True)
    app.run(host="0.0.0.0", port=port, debug=False)
